import openpyxl
from openpyxl.styles import Font, Alignment
import os
import mysql.connector

db_host = "localhost"
db_username = "root"
db_password = ""
db_name = "db_ecook_data"

filename = "ecook_data_comparison.xlsx"

if os.path.exists(filename):
    os.remove(filename)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comparison"

A1 = ws.cell(row=1, column=1)
B1 = ws.cell(row=1, column=2)
C1 = ws.cell(row=1, column=3)
D1 = ws.cell(row=1, column=4)

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 15


ws['A1'].font = Font(size= 12, bold=True)
ws['B1'].font = Font(size= 12, bold=True)
ws['C1'].font = Font(size= 12, bold=True)
ws['D1'].font = Font(size= 12, bold=True)


A1.value = "Date"
B1.value = "Number of Active Units"
C1.value = "Units Registered till Date"
D1.value = "Active (%)"

row_counter = 2


query_string = "SELECT reports_dailyusagedata.when_date AS 'DATE', COUNT(reports_dailyusagedata.serial_number) AS 'Number of Active Units', (SELECT COUNT(tbl_ecook.unit_number) FROM tbl_ecook WHERE tbl_ecook.registration_date <= when_date ) AS 'Number of Unit Registered till Date' FROM reports_dailyusagedata GROUP BY when_date ORDER BY reports_dailyusagedata.when_date ASC"

db_con = mysql.connector.connect(host=db_host, user=db_username, password=db_password, database=db_name)

db_cursor = db_con.cursor()
db_cursor.execute(query_string)
db_result = db_cursor.fetchall()

for x in db_result:
    cell_1 = ws.cell(row=row_counter, column=1)
    cell_2 = ws.cell(row=row_counter, column=2)
    cell_3 = ws.cell(row=row_counter, column=3)
    cell_4 = ws.cell(row=row_counter, column=4)
    
    cell_1.value = x[0]
    cell_2.value = x[1]
    cell_3.value = x[2]
    cell_4.value = "{:.2f}".format((x[1] / x[2]) * 100) + "%"
    cell_4.alignment = Alignment(horizontal='right')

    row_counter += 1

wb.save(filename)