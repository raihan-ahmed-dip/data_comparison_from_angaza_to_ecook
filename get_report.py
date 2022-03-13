import openpyxl
from openpyxl.styles import Font, Alignment
import os
import mysql.connector
import datetime

start_date = datetime.date(2022, 1, 1)
end_date = datetime.date.today() - datetime.timedelta(days=1)
delta = datetime.timedelta(days=1)

db_host = "localhost"
db_username = "root"
db_password = ""
db_name = "db_ecook_data"
db_con = mysql.connector.connect(host=db_host, user=db_username, password=db_password, database=db_name)


filename = "ecook_data_comparison.xlsx"

if os.path.exists(filename):
    os.remove(filename)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comparison"

A1 = ws.cell(row=1, column=1)
B1 = ws.cell(row=1, column=2)
C1 = ws.cell(row=1, column=3)
D1 = ws.cell(row=1, column=4)
E1 = ws.cell(row=1, column=5)
F1 = ws.cell(row=1, column=6)
G1 = ws.cell(row=1, column=7)
H1 = ws.cell(row=1, column=8)
I1 = ws.cell(row=1, column=9)
J1 = ws.cell(row=1, column=10)
K1 = ws.cell(row=1, column=11)
L1 = ws.cell(row=1, column=12)
M1 = ws.cell(row=1, column=13)
N1 = ws.cell(row=1, column=14)
O1 = ws.cell(row=1, column=15)
P1 = ws.cell(row=1, column=16)
Q1 = ws.cell(row=1, column=17)
R1 = ws.cell(row=1, column=18)
S1 = ws.cell(row=1, column=19)
T1 = ws.cell(row=1, column=20)
U1 = ws.cell(row=1, column=21)
V1 = ws.cell(row=1, column=22)

A1.value = "Date"
B1.value = "BD Registered eCook"
C1.value = "BD Active ecook"
D1.value = "BD Active (%)"
E1.value = "CAM Registered eCook"
F1.value = "CAM Active ecook"
G1.value = "CAM Active (%)"
H1.value = "Double 2000W Registered eCook"
I1.value = "Double 2000W Active ecook"
J1.value = "Double 2000W Active (%)"
K1.value = "Double 2000W v2 Registered eCook"
L1.value = "Double 2000W v2 Active ecook"
M1.value = "Double 2000W v2 Active (%)"
N1.value = "Single 2000W Registered eCook"
O1.value = "Single 2000W Active ecook"
P1.value = "Single 2000W Active (%)"
Q1.value = "Single 1600W Registered eCook"
R1.value = "Single 1600W Active ecook"
S1.value = "Single 1600W Active (%)"
T1.value = "Total Registered eCook"
U1.value = "Total Active ecook"
V1.value = "Total Active (%)"

row_counter = 2

while start_date <= end_date:
    cell_1 = ws.cell(row=row_counter, column=1)
    cell_2 = ws.cell(row=row_counter, column=2)
    cell_3 = ws.cell(row=row_counter, column=3)
    cell_4 = ws.cell(row=row_counter, column=4)
    cell_5 = ws.cell(row=row_counter, column=5)
    cell_6 = ws.cell(row=row_counter, column=6)
    cell_7 = ws.cell(row=row_counter, column=7)
    cell_8 = ws.cell(row=row_counter, column=8)
    cell_9 = ws.cell(row=row_counter, column=9)
    cell_10 = ws.cell(row=row_counter, column=10)
    cell_11 = ws.cell(row=row_counter, column=11)
    cell_12 = ws.cell(row=row_counter, column=12)
    cell_13 = ws.cell(row=row_counter, column=13)
    cell_14 = ws.cell(row=row_counter, column=14)
    cell_15 = ws.cell(row=row_counter, column=15)
    cell_16 = ws.cell(row=row_counter, column=16)
    cell_17 = ws.cell(row=row_counter, column=17)
    cell_18 = ws.cell(row=row_counter, column=18)
    cell_19 = ws.cell(row=row_counter, column=19)
    cell_20 = ws.cell(row=row_counter, column=20)
    cell_21 = ws.cell(row=row_counter, column=21)
    cell_22 = ws.cell(row=row_counter, column=22)
    
    date = start_date
    
    bd_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'BD Units' FROM tbl_ecook WHERE tbl_ecook.country = 'Bangladesh' AND tbl_ecook.registration_date <= '{date}'"
    bd_eook_registered_cursor = db_con.cursor()
    bd_eook_registered_cursor.execute(bd_registered_query)
    bd_registered_ecook = bd_eook_registered_cursor.fetchone()[0]

    bd_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE tbl_ecook.country= 'Bangladesh' AND reports_dailyusagedata.when_date = '{date}'"
    bd_active_ecook_cursor = db_con.cursor()
    bd_eook_registered_cursor.execute(bd_active_query)
    bd_active_ecook = bd_eook_registered_cursor.fetchone()[0]

    cam_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'CAM Units' FROM tbl_ecook WHERE tbl_ecook.country = 'Cambodia' AND tbl_ecook.registration_date <= '{date}'"
    cam_eook_registered_cursor = db_con.cursor()
    cam_eook_registered_cursor.execute(cam_registered_query)
    cam_registered_ecook = cam_eook_registered_cursor.fetchone()[0]
    
    cam_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE tbl_ecook.country= 'Cambodia' AND reports_dailyusagedata.when_date = '{date}'"
    cam_active_ecook_cursor = db_con.cursor()
    cam_eook_registered_cursor.execute(cam_active_query)
    cam_active_ecook = cam_eook_registered_cursor.fetchone()[0]

    double_v1_2000_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'CAM Units' FROM tbl_ecook WHERE tbl_ecook.product = 'ATEC GSM Double Stove 2000W' AND tbl_ecook.registration_date <= '{date}'"
    double_v1_2000_eook_registered_cursor = db_con.cursor()
    double_v1_2000_eook_registered_cursor.execute(double_v1_2000_registered_query)
    double_v1_2000_registered_ecook = double_v1_2000_eook_registered_cursor.fetchone()[0]

    double_v1_2000_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE tbl_ecook.product = 'ATEC GSM Double Stove 2000W' AND reports_dailyusagedata.when_date = '{date}'"
    double_v1_2000_active_ecook_cursor = db_con.cursor()
    double_v1_2000_eook_registered_cursor.execute(double_v1_2000_active_query)
    double_v1_2000_active_ecook = double_v1_2000_eook_registered_cursor.fetchone()[0]

    single_v1_1600_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'CAM Units' FROM tbl_ecook WHERE tbl_ecook.product = 'ATEC GSM Single Stove 1600W' AND tbl_ecook.registration_date <= '{date}'"
    single_v1_1600_eook_registered_cursor = db_con.cursor()
    single_v1_1600_eook_registered_cursor.execute(single_v1_1600_registered_query)
    single_v1_1600_registered_ecook = single_v1_1600_eook_registered_cursor.fetchone()[0]

    single_v1_1600_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE tbl_ecook.product = 'ATEC GSM Single Stove 1600W' AND reports_dailyusagedata.when_date = '{date}'"
    single_v1_1600_active_ecook_cursor = db_con.cursor()
    single_v1_1600_eook_registered_cursor.execute(single_v1_1600_active_query)
    single_v1_1600_active_ecook = single_v1_1600_eook_registered_cursor.fetchone()[0]

    single_v1_2000_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'CAM Units' FROM tbl_ecook WHERE tbl_ecook.product = 'ATEC GSM Single Stove 2000W' AND tbl_ecook.registration_date <= '{date}'"
    single_v1_2000_eook_registered_cursor = db_con.cursor()
    single_v1_2000_eook_registered_cursor.execute(single_v1_2000_registered_query)
    single_v1_2000_registered_ecook = single_v1_2000_eook_registered_cursor.fetchone()[0]

    single_v1_2000_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE tbl_ecook.product = 'ATEC GSM Single Stove 2000W' AND reports_dailyusagedata.when_date = '{date}'"
    single_v1_2000_active_ecook_cursor = db_con.cursor()
    single_v1_2000_eook_registered_cursor.execute(single_v1_2000_active_query)
    single_v1_2000_active_ecook = single_v1_2000_eook_registered_cursor.fetchone()[0]

    double_v2_2000_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'CAM Units' FROM tbl_ecook WHERE tbl_ecook.product = 'GSM Double Stove 2000W v2' AND tbl_ecook.registration_date <= '{date}'"
    double_v2_2000_eook_registered_cursor = db_con.cursor()
    double_v2_2000_eook_registered_cursor.execute(double_v2_2000_registered_query)
    double_v2_2000_registered_ecook = double_v2_2000_eook_registered_cursor.fetchone()[0]

    double_v2_2000_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE tbl_ecook.product = 'GSM Double Stove 2000W v2' AND reports_dailyusagedata.when_date = '{date}'"
    double_v2_2000_active_ecook_cursor = db_con.cursor()
    double_v2_2000_eook_registered_cursor.execute(double_v2_2000_active_query)
    double_v2_2000_active_ecook = double_v2_2000_eook_registered_cursor.fetchone()[0]

    total_registered_query = f"SELECT COUNT(tbl_ecook.unit_number) AS 'BD Units' FROM tbl_ecook WHERE tbl_ecook.registration_date <= '{date}'"
    total_eook_registered_cursor = db_con.cursor()
    total_eook_registered_cursor.execute(total_registered_query)
    total_registered_ecook = total_eook_registered_cursor.fetchone()[0]

    total_active_query = f"SELECT COUNT(reports_dailyusagedata.serial_number) FROM reports_dailyusagedata INNER JOIN tbl_ecook ON reports_dailyusagedata.serial_number = tbl_ecook.unit_number WHERE reports_dailyusagedata.when_date = '{date}'"
    total_active_ecook_cursor = db_con.cursor()
    total_eook_registered_cursor.execute(total_active_query)
    total_active_ecook = total_eook_registered_cursor.fetchone()[0]

    cell_1.value = date
    cell_2.value = bd_registered_ecook
    cell_3.value = bd_active_ecook
    cell_4.value = "{:.2f}".format((bd_active_ecook / bd_registered_ecook) * 100) + "%"
    cell_5.value = cam_registered_ecook
    cell_6.value = cam_active_ecook
    cell_7.value = "{:.2f}".format((cam_active_ecook / cam_registered_ecook) * 100) + "%"
    cell_8.value = double_v1_2000_registered_ecook
    cell_9.value = double_v1_2000_active_ecook
    cell_10.value = "{:.2f}".format((double_v1_2000_active_ecook / double_v1_2000_registered_ecook) * 100) + "%"
    cell_11.value = double_v2_2000_registered_ecook
    cell_12.value = double_v2_2000_active_ecook
    cell_13.value = "{:.2f}".format((double_v2_2000_active_ecook / double_v2_2000_registered_ecook) * 100) + "%"
    cell_14.value = single_v1_2000_registered_ecook
    cell_15.value = single_v1_2000_active_ecook
    cell_16.value = "{:.2f}".format((single_v1_2000_active_ecook / single_v1_2000_registered_ecook) * 100) + "%"
    cell_17.value = single_v1_1600_registered_ecook
    cell_18.value = single_v1_1600_active_ecook
    cell_19.value = "{:.2f}".format((single_v1_1600_active_ecook / single_v1_1600_registered_ecook) * 100) + "%"
    cell_20.value = total_registered_ecook
    cell_21.value = total_active_ecook
    cell_22.value = "{:.2f}".format((total_active_ecook / total_registered_ecook) * 100) + "%"

    start_date += delta
    row_counter += 1
    print(f"{date} :: {bd_registered_ecook} :: {bd_active_ecook} :: {cam_registered_ecook} :: {cam_active_ecook} :: {double_v1_2000_registered_ecook} :: {double_v1_2000_active_ecook} :: {single_v1_1600_registered_ecook} :: {single_v1_1600_active_ecook} :: {single_v1_2000_registered_ecook} :: {single_v1_2000_active_ecook} :: {double_v2_2000_registered_ecook} :: {double_v2_2000_active_ecook}")

wb.save(filename)